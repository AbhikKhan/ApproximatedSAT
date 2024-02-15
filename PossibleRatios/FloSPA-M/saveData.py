import pandas as pd

import pandas as pd
from openpyxl import load_workbook
import pandas as pd
from openpyxl import load_workbook

def save_data_to_excel(data, file_path):
    # Define the column order
    column_order = [{'Ratio', 'area', 'Loading Cycles (K)', 'Bendings (B)', 'Path Length (L)'}]
    # Create a DataFrame object with specified column order
    df = pd.DataFrame(data, columns=column_order)
    
    # Load the existing Excel file
    book = load_workbook(file_path)
    writer = pd.ExcelWriter(file_path, engine='openpyxl') 
    writer.book = book
    
    # Append the data to the existing sheet or create a new sheet
    sheet_name = 'Sheet1'  # Change to the desired sheet name
    try:
        writer.sheets = {ws.title: ws for ws in book.worksheets}
        df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=writer.sheets[sheet_name].max_row)
    except KeyError:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    # Save the changes
    writer.save()
    writer.close()